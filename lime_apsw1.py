import apsw
import json
from pprint import pprint
import random
import datetime
import openpyxl
import sys

# DB_FILE='db_apsw.sqlite3'
DB_NAME = 'example3.sqlite3'


def create_table():
    """
    tableを作成する
    """

    sql = """
        create table tbl1(
            id int not null primary key,
            name text,
            age int,
            weight real,
            json json
        )
    """

    connection = apsw.Connection(DB_NAME)
    cursor = connection.cursor()
    cursor.execute(sql)


def insert_data():
    """
    データを一件、挿入する
    """

    dctData = {}
    dctData['名前'] = '斉藤'
    dctData['年齢'] = 25
    dctData['体重'] = 54.3
    dctData['入社日'] = '1995-09-15'
    json_text = json.dumps(dctData, ensure_ascii=False)

    connection = apsw.Connection(DB_NAME)
    cursor = connection.cursor()
    cursor.execute("insert into tbl1 values(?,?,?,?,?)",
                   (2, '鹿児島', 23, 54.3, json_text))


def insert_data_mary():
    """
    データを配列でまとめて、挿入する
    """

    dctData = {}
    dctData['名前'] = '斉藤'
    dctData['年齢'] = 25
    dctData['体重'] = 54.3
    dctData['入社日'] = '1995-09-15'
    json_text = json.dumps(dctData, ensure_ascii=False)

    lstData = [
        [3, '東京', 23, 44.3, json_text],
        [4, '大阪', 53, 64.3, json_text],
        [5, '福岡', 63, 74.3, json_text],
    ]

    connection = apsw.Connection(DB_NAME)
    cursor = connection.cursor()
    cursor.executemany("insert into tbl1 values(?,?,?,?,?)", lstData)


def insert_dict():
    """
    辞書を一件、挿入する
    """

    dctData = {}
    dctData['名前'] = '斉藤'
    dctData['年齢'] = 25
    dctData['体重'] = 54.3
    dctData['入社日'] = '1995-09-15'
    json_text = json.dumps(dctData, ensure_ascii=False)

    dctTemp = {}
    dctTemp['id'] = 6
    dctTemp['name'] = '北九州市'
    dctTemp['age'] = 21
    dctTemp['weight'] = 53.9
    dctTemp['json'] = json_text

    connection = apsw.Connection(DB_NAME)
    cursor = connection.cursor()
    cursor.executemany("insert into t_lime2 values(?,?)", lstData)


def insert_dict_mary(lstData):
    """
    辞書を配列でまとめて、挿入する
    """

    connection = apsw.Connection(DB_NAME)
    cursor = connection.cursor()
    cursor.execute("BEGIN TRANSACTION;")

    for row in lstData:
        cursor.execute("insert into t_lime2 values(?,?)", row)

    cursor.execute("COMMIT;")


def select_rows():
    """
    データ select
    """

    connection = apsw.Connection(DB_NAME)
    cursor = connection.cursor()
    cmd = "SELECT * FROM tbl1 WHERE age>50 "

    for row in cursor.execute(cmd):
        pprint(row)


def select_rows_json():
    """
    データ select (json内の特定keyで検索)
    """

    connection = apsw.Connection(DB_NAME)
    cursor = connection.cursor()
    # cmd = "SELECT * FROM tbl1 WHERE json_extract(tbl1.json, '$.名前') like '斉藤%' "
    # cmd = "SELECT * FROM tbl1 WHERE json_extract(tbl1.json, '$.年齢')>50 "
    cmd = ("SELECT * FROM tbl1 WHERE"
           " json_extract(tbl1.json, '$.入社日')>'2020-01-01' ")

    for row in cursor.execute(cmd):
        json_text = row[4]
        dctData = json.loads(json_text)  # jsonを辞書に変換
        pprint(dctData)


def read_excel():
    """
    既存のexcelを読み込む
    """

    excel_file = '郵便番号.xlsx'
    # excel_file = '郵便番号_min.xlsx'
    sheet_name = 'yubin_all'

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.get_sheet_by_name(sheet_name)

    # for col in range(1,  ws.max_column + 1):
    #     name = ws.cell(row=2, column=col).value
    #     print('{}-{}'.format(col, name))

    max_col = ws.max_column
    lstData = []
    for row in range(3,  ws.max_row + 1):
        # if row > 10:
        #     break
        print(row)

        dctTemp = {}
        for col in range(1,  max_col + 1):
            val = ws.cell(row=row, column=col).value
            key = 'A{:04d}'.format(col)
            # print('{}-{}-{}-{}'.format(row, col, val, key))
            dctTemp[key] = val
        # pprint(dctTemp)

        id = ws.cell(row=row, column=1).value
        lstData.append([id, json.dumps(dctTemp, ensure_ascii=False)])

    # pprint(lstData)

    return lstData


def read_excel2():
    """
    既存のexcelを読み込む
    """

    excel_file = '薬価3_min.xlsx'
    # excel_file = '薬価3.xlsx'
    sheet_name = '薬価3'

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.get_sheet_by_name(sheet_name)

    # for col in range(1,  ws.max_column + 1):
    #     name = ws.cell(row=2, column=col).value
    #     print('{}-{}'.format(col, name))

    max_col = ws.max_column
    lstData = []
    for row in range(3,  ws.max_row + 1):
        # if row > 10:
        #     break
        # print(row)

        dctTemp = {}
        for col in range(1,  max_col + 1):
            val = ws.cell(row=row, column=col).value
            key = 'A{:04d}'.format(col)
            # print('{}-{}'.format(val,type(val)))
            if type(val) is datetime.datetime:
                dctTemp[key] = val.strftime("%Y-%m-%d")
            else:
                dctTemp[key] = val

        # pprint(dctTemp)

        id = ws.cell(row=row, column=1).value
        lstData.append([id, json.dumps(dctTemp, ensure_ascii=False)])

    pprint(lstData)

    return lstData


def read_def(tbl_id):
    """
    excelの定義ファイルを読み込む
    """

    excel_file = '薬価3_定義.xlsx'
    sheet_name = '薬価3'

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.get_sheet_by_name(sheet_name)

    max_col = ws.max_column

    lstData = []
    for col in range(1,  max_col + 1):
        col_id = col
        col_name = 'c{:03}{:03}'.format(tbl_id, col_id)
        col_level1 = ws.cell(row=1, column=col).value
        col_level2 = ws.cell(row=2, column=col).value
        col_level3 = ws.cell(row=3, column=col).value
        col_type = ws.cell(row=4, column=col).value
        col_format = ws.cell(row=5, column=col).value
        lstData.append([tbl_id, col_id, col_name, col_level1, col_level2,
                        col_level3, col_type, col_format])

    # pprint(lstData)

    return lstData


def create_apl_table():
    """
    apl_tableを作成する
    """

    connection = apsw.Connection(DB_NAME)
    cursor = connection.cursor()

    for cnt in range(10):

        sql = """
            create table apl_{:03}(
                id text not null primary key,
                attr json
            )
        """.format(cnt)

        print(sql)
        cursor.execute(sql)


def insert_dict_mary2(lstData):
    """
    辞書を配列でまとめて、挿入する
    """

    connection = apsw.Connection(DB_NAME)
    cursor = connection.cursor()
    cursor.execute("BEGIN TRANSACTION;")

    for row in lstData:
        cursor.execute("insert into apl_000 values(?,?)", row)

    cursor.execute("COMMIT;")


def insert_m_colomn(lstData):
    """

    """

    connection = apsw.Connection(DB_NAME)
    cursor = connection.cursor()
    cursor.execute("BEGIN TRANSACTION;")

    for row in lstData:
        cursor.execute(
            "insert into m_column values(?, ?, ?, ?, ?, ?, ?, ?)", row)

    cursor.execute("COMMIT;")


if __name__ == '__main__':

    # create_apl_table()

    lstData = read_def(0)
    insert_m_colomn(lstData)
    sys.exit()

    # lstData = read_excel()

    # lstData = read_excel2()
    # pprint('loaded excel')
    # insert_dict_mary2(lstData)
    # pprint('finished')
    # sys.exit()

    # a = [3, 4, 5, 5]
    # base_date=datetime.date(2017, 11, 12)
    # for cnt in range(70,80):
    #     day=base_date + datetime.timedelta(days=random.randint(0,1000))
    #     print(base_date)
    #     print(day)

    # sys.exit()

    # select_data()

    # create_table()
    # insert_data()
    # insert_data_mary()
    # insert_dict()

    # insert_dict_mary()
    # select_rows_json()
